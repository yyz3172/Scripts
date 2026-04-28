#!/usr/bin/env python3
"""
从各 batch 的 LongBench 输出 jsonl 中汇总准确率与延迟指标，输出一个 Excel 文件。

参考：
- `parse_aisbench_log.py` 的目录/CLI 结构与 Excel 风格：传入父目录，自动收集 batch_* 子目录。
- `code/LongBench/result.py` 的统计口径：overall / easy / hard / short / medium / long，
  以及 ttft/e2e/tpot（含 cot 二段）延迟分布统计（avg/min/max/median/p90/p99）。

输入：
- <log_dir>/batch_*/long_bench_output.jsonl（文件名固定）

输出：
- <log_dir>/longbench_summary.xlsx（可用 --output 改名）
"""

from __future__ import annotations

import argparse
import json
import math
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


DEFAULT_FILENAME = "long_bench_output.jsonl"
DEFAULT_OUTPUT_XLSX = "longbench_summary.xlsx"
PRINT_SEP_WIDTH = 140

# -----------------------------------------------------------------------------
# Excel 样式/图表配置（尽量与 parse_aisbench_log.py 对齐）
# -----------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
COL_WIDTH_INDEX = 34
COL_WIDTH_DATA = 20
CHART_WIDTH = 40
CHART_HEIGHT = 20
CHART_AXIS_MARGIN_RATIO = 0.05
CHART_AXIS_MARGIN_RATIO_MIN = 0.02
SHEET_TITLE_MAX_LEN = 31

SUMMARY_SHEET_TITLE = "汇总"
SUMMARY_HEADER_LABEL = "指标"
SUMMARY_SUBLABEL = "子项"

CHART_GROUPS = [
    ("Accuracy", ["acc_overall", "acc_easy", "acc_hard", "acc_short", "acc_medium", "acc_long"]),
    ("TTFT(ms)", ["ttft_n", "ttft_avg", "ttft_max", "ttft_p99", "ttft_p95", "ttft_p90", "ttft_p50", "ttft_min"]),
    ("E2E(ms)", ["e2e_n", "e2e_avg", "e2e_max", "e2e_p99", "e2e_p95", "e2e_p90", "e2e_p50", "e2e_min"]),
    ("TPOT(ms)", ["tpot_n", "tpot_avg", "tpot_max", "tpot_p99", "tpot_p95", "tpot_p90", "tpot_p50", "tpot_min"]),
    ("InputTokens(tok)", ["input_tokens_n", "input_tokens_avg", "input_tokens_max", "input_tokens_p99", "input_tokens_p95", "input_tokens_p90", "input_tokens_p50", "input_tokens_min"]),
    ("OutputTokens(tok)", ["output_tokens_n", "output_tokens_avg", "output_tokens_max", "output_tokens_p99", "output_tokens_p95", "output_tokens_p90", "output_tokens_p50", "output_tokens_min"]),
    ("InputTokens_COT(tok)", ["input_tokens_cot_n", "input_tokens_cot_avg", "input_tokens_cot_max", "input_tokens_cot_p99", "input_tokens_cot_p95", "input_tokens_cot_p90", "input_tokens_cot_p50", "input_tokens_cot_min"]),
    ("OutputTokens_COT(tok)", ["output_tokens_cot_n", "output_tokens_cot_avg", "output_tokens_cot_max", "output_tokens_cot_p99", "output_tokens_cot_p95", "output_tokens_cot_p90", "output_tokens_cot_p50", "output_tokens_cot_min"]),
    ("TTFT_COT(ms)", ["ttft_cot_n", "ttft_cot_avg", "ttft_cot_max", "ttft_cot_p99", "ttft_cot_p95", "ttft_cot_p90", "ttft_cot_p50", "ttft_cot_min"]),
    ("E2E_COT(ms)", ["e2e_cot_n", "e2e_cot_avg", "e2e_cot_max", "e2e_cot_p99", "e2e_cot_p95", "e2e_cot_p90", "e2e_cot_p50", "e2e_cot_min"]),
]
# -----------------------------------------------------------------------------


def _read_jsonl(path: str) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            out.append(json.loads(line))
    return out


def _pct(numer: float, denom: int) -> Optional[float]:
    if denom == 0:
        return None
    return 100.0 * float(numer) / float(denom)


def _percentile(sorted_vals: List[float], p: float) -> Optional[float]:
    """
    p: 0-100
    使用最近秩（nearest-rank）定义：ceil(p/100*N) 的值（与 `LongBench/result.py` 对齐）。
    """
    if not sorted_vals:
        return None
    if p <= 0:
        return float(sorted_vals[0])
    if p >= 100:
        return float(sorted_vals[-1])
    n = len(sorted_vals)
    k = int(math.ceil(p / 100.0 * n)) - 1
    k = max(0, min(n - 1, k))
    return float(sorted_vals[k])


def _summarize_latency(vals: List[Any]) -> Dict[str, Any]:
    v2 = []
    for v in vals:
        if isinstance(v, (int, float)) and not math.isnan(v):
            v2.append(float(v))
    v2.sort()
    if not v2:
        return {
            "n": 0,
            "avg": None,
            "min": None,
            "max": None,
            "p50": None,
            "p95": None,
            "p90": None,
            "p99": None,
        }
    n = len(v2)
    avg = sum(v2) / n
    return {
        "n": n,
        "avg": avg,
        "min": float(v2[0]),
        "max": float(v2[-1]),
        "p50": _percentile(v2, 50),
        "p95": _percentile(v2, 95),
        "p90": _percentile(v2, 90),
        "p99": _percentile(v2, 99),
    }


@dataclass
class BatchStats:
    batch: str
    file: str
    n: int
    pred_none: int
    accuracy: Dict[str, Optional[float]]
    latency_ms: Dict[str, Dict[str, Any]]
    # 仅针对 TTFT/E2E/TPOT/OutputTokens，按 domain + sub_domain 分组统计
    # key: "<domain>\t<sub_domain>"
    latency_by_domain_sub: Dict[str, Dict[str, Dict[str, Any]]]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "batch": self.batch,
            "file": self.file,
            "n": self.n,
            "pred_none": self.pred_none,
            "accuracy": self.accuracy,
            "latency_ms": self.latency_ms,
            "latency_by_domain_sub": self.latency_by_domain_sub,
        }


def collect_batches(log_dir: str) -> List[str]:
    """在 log_dir 下查找 batch_* 子目录（含固定文件名），按目录名排序。"""
    names: List[str] = []
    for name in os.listdir(log_dir):
        full = os.path.join(log_dir, name)
        if not name.startswith("batch_") or not os.path.isdir(full):
            continue
        if not os.path.isfile(os.path.join(full, DEFAULT_FILENAME)):
            continue
        names.append(name)
    names.sort()
    return names


def summarize_one(jsonl_path: str, batch_name: str) -> BatchStats:
    pred_data = _read_jsonl(jsonl_path)
    n = len(pred_data)

    easy = hard = short = medium = long_ = 0
    easy_acc = hard_acc = short_acc = medium_acc = long_acc = 0.0

    ttft_ms: List[Any] = []
    e2e_ms: List[Any] = []
    tpot_ms: List[Any] = []
    input_tokens: List[Any] = []
    output_tokens: List[Any] = []
    ttft_ms_cot: List[Any] = []
    e2e_ms_cot: List[Any] = []
    input_tokens_cot: List[Any] = []
    output_tokens_cot: List[Any] = []

    pred_none = 0
    # TTFT/E2E/TPOT/OutputTokens 按 domain/sub_domain 分组
    group_ttft: Dict[tuple[str, str], List[Any]] = {}
    group_e2e: Dict[tuple[str, str], List[Any]] = {}
    group_tpot: Dict[tuple[str, str], List[Any]] = {}
    group_input_tokens: Dict[tuple[str, str], List[Any]] = {}
    group_output_tokens: Dict[tuple[str, str], List[Any]] = {}
    group_input_tokens_cot: Dict[tuple[str, str], List[Any]] = {}
    group_output_tokens_cot: Dict[tuple[str, str], List[Any]] = {}

    for pred in pred_data:
        acc = float(bool(pred.get("judge")))
        if pred.get("pred") is None:
            pred_none += 1

        if pred.get("difficulty") == "easy":
            easy += 1
            easy_acc += acc
        else:
            hard += 1
            hard_acc += acc

        length = pred.get("length")
        if length == "short":
            short += 1
            short_acc += acc
        elif length == "medium":
            medium += 1
            medium_acc += acc
        else:
            long_ += 1
            long_acc += acc

        if "ttft_ms" in pred:
            ttft_ms.append(pred.get("ttft_ms"))
        if "e2e_ms" in pred:
            e2e_ms.append(pred.get("e2e_ms"))
        if "tpot_ms" in pred:
            tpot_ms.append(pred.get("tpot_ms"))
        if "input_tokens" in pred:
            input_tokens.append(pred.get("input_tokens"))
        if "output_tokens" in pred:
            output_tokens.append(pred.get("output_tokens"))
        if "ttft_ms_cot" in pred:
            ttft_ms_cot.append(pred.get("ttft_ms_cot"))
        if "e2e_ms_cot" in pred:
            e2e_ms_cot.append(pred.get("e2e_ms_cot"))
        if "input_tokens_cot" in pred:
            input_tokens_cot.append(pred.get("input_tokens_cot"))
        if "output_tokens_cot" in pred:
            output_tokens_cot.append(pred.get("output_tokens_cot"))

        dom = str(pred.get("domain") or "")
        sub = str(pred.get("sub_domain") or "")
        gk = (dom, sub)
        if "ttft_ms" in pred:
            group_ttft.setdefault(gk, []).append(pred.get("ttft_ms"))
            group_ttft.setdefault((dom, "all"), []).append(pred.get("ttft_ms"))
            group_ttft.setdefault(("all", "all"), []).append(pred.get("ttft_ms"))
        if "e2e_ms" in pred:
            group_e2e.setdefault(gk, []).append(pred.get("e2e_ms"))
            group_e2e.setdefault((dom, "all"), []).append(pred.get("e2e_ms"))
            group_e2e.setdefault(("all", "all"), []).append(pred.get("e2e_ms"))
        if "tpot_ms" in pred:
            group_tpot.setdefault(gk, []).append(pred.get("tpot_ms"))
            group_tpot.setdefault((dom, "all"), []).append(pred.get("tpot_ms"))
            group_tpot.setdefault(("all", "all"), []).append(pred.get("tpot_ms"))
        if "input_tokens" in pred:
            group_input_tokens.setdefault(gk, []).append(pred.get("input_tokens"))
            group_input_tokens.setdefault((dom, "all"), []).append(pred.get("input_tokens"))
            group_input_tokens.setdefault(("all", "all"), []).append(pred.get("input_tokens"))
        if "output_tokens" in pred:
            group_output_tokens.setdefault(gk, []).append(pred.get("output_tokens"))
            group_output_tokens.setdefault((dom, "all"), []).append(pred.get("output_tokens"))
            group_output_tokens.setdefault(("all", "all"), []).append(pred.get("output_tokens"))
        if "input_tokens_cot" in pred:
            group_input_tokens_cot.setdefault(gk, []).append(pred.get("input_tokens_cot"))
            group_input_tokens_cot.setdefault((dom, "all"), []).append(pred.get("input_tokens_cot"))
            group_input_tokens_cot.setdefault(("all", "all"), []).append(pred.get("input_tokens_cot"))
        if "output_tokens_cot" in pred:
            group_output_tokens_cot.setdefault(gk, []).append(pred.get("output_tokens_cot"))
            group_output_tokens_cot.setdefault((dom, "all"), []).append(pred.get("output_tokens_cot"))
            group_output_tokens_cot.setdefault(("all", "all"), []).append(pred.get("output_tokens_cot"))

    overall = _pct(easy_acc + hard_acc, n)
    acc = {
        "overall": overall,
        "easy": _pct(easy_acc, easy),
        "hard": _pct(hard_acc, hard),
        "short": _pct(short_acc, short),
        "medium": _pct(medium_acc, medium),
        "long": _pct(long_acc, long_),
    }
    latency = {
        "ttft": _summarize_latency(ttft_ms),
        "e2e": _summarize_latency(e2e_ms),
        "tpot": _summarize_latency(tpot_ms),
        "input_tokens": _summarize_latency(input_tokens),
        "output_tokens": _summarize_latency(output_tokens),
        "ttft_cot": _summarize_latency(ttft_ms_cot),
        "e2e_cot": _summarize_latency(e2e_ms_cot),
        "input_tokens_cot": _summarize_latency(input_tokens_cot),
        "output_tokens_cot": _summarize_latency(output_tokens_cot),
    }

    latency_by_domain_sub: Dict[str, Dict[str, Dict[str, Any]]] = {}
    all_groups = (
        set(group_ttft.keys())
        | set(group_e2e.keys())
        | set(group_tpot.keys())
        | set(group_input_tokens.keys())
        | set(group_output_tokens.keys())
        | set(group_input_tokens_cot.keys())
        | set(group_output_tokens_cot.keys())
    )

    def _gk_sort(x: tuple[str, str]) -> tuple[int, str, int, str]:
        d, s = x
        # all 放最前；sub_domain=all 也优先
        return (0 if d == "all" else 1, d, 0 if s == "all" else 1, s)

    for dom, sub in sorted(all_groups, key=_gk_sort):
        kk = f"{dom}\t{sub}"
        latency_by_domain_sub[kk] = {
            "TTFT": _summarize_latency(group_ttft.get((dom, sub), [])),
            "E2E": _summarize_latency(group_e2e.get((dom, sub), [])),
            "TPOT": _summarize_latency(group_tpot.get((dom, sub), [])),
            "InputTokens": _summarize_latency(group_input_tokens.get((dom, sub), [])),
            "OutputTokens": _summarize_latency(group_output_tokens.get((dom, sub), [])),
            "InputTokens_COT": _summarize_latency(group_input_tokens_cot.get((dom, sub), [])),
            "OutputTokens_COT": _summarize_latency(group_output_tokens_cot.get((dom, sub), [])),
        }

    return BatchStats(
        batch=batch_name,
        file=jsonl_path,
        n=n,
        pred_none=pred_none,
        accuracy=acc,
        latency_ms=latency,
        latency_by_domain_sub=latency_by_domain_sub,
    )


def _fmt(v: Any, nd: int = 1) -> str:
    if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
        return f"{float(v):.{nd}f}"
    return "--"


def _flatten_stats(s: BatchStats) -> Dict[str, Any]:
    """展开成用于 Excel 的扁平 key->value。"""
    out: Dict[str, Any] = {
        "batch": s.batch,
        "n": s.n,
        "pred_none": s.pred_none,
        "acc_overall": s.accuracy.get("overall"),
        "acc_easy": s.accuracy.get("easy"),
        "acc_hard": s.accuracy.get("hard"),
        "acc_short": s.accuracy.get("short"),
        "acc_medium": s.accuracy.get("medium"),
        "acc_long": s.accuracy.get("long"),
    }

    def _put(prefix: str, d: Dict[str, Any]):
        out[f"{prefix}_n"] = d.get("n")
        out[f"{prefix}_avg"] = d.get("avg")
        out[f"{prefix}_min"] = d.get("min")
        out[f"{prefix}_max"] = d.get("max")
        out[f"{prefix}_p50"] = d.get("p50")
        out[f"{prefix}_p95"] = d.get("p95")
        out[f"{prefix}_p90"] = d.get("p90")
        out[f"{prefix}_p99"] = d.get("p99")

    _put("ttft", s.latency_ms.get("ttft", {}))
    _put("e2e", s.latency_ms.get("e2e", {}))
    _put("tpot", s.latency_ms.get("tpot", {}))
    _put("input_tokens", s.latency_ms.get("input_tokens", {}))
    _put("output_tokens", s.latency_ms.get("output_tokens", {}))
    _put("ttft_cot", s.latency_ms.get("ttft_cot", {}))
    _put("e2e_cot", s.latency_ms.get("e2e_cot", {}))
    _put("input_tokens_cot", s.latency_ms.get("input_tokens_cot", {}))
    _put("output_tokens_cot", s.latency_ms.get("output_tokens_cot", {}))
    return out


def _metric_unit(key: str) -> str:
    if key.startswith("acc_"):
        return "%"
    if key in ("n", "pred_none"):
        return ""
    if key.startswith(("input_tokens_", "output_tokens_", "input_tokens_cot_", "output_tokens_cot_")):
        return "tok"
    if key.endswith(("_avg", "_min", "_max", "_p50", "_p90", "_p95", "_p99")):
        # latency in ms
        return "ms"
    return ""


def _ordered_latency_keys(prefix: str) -> List[str]:
    """
    同一数据项的指标顺序：
    n、avg、max、P99、P95、P90、P50、min
    """
    return [
        f"{prefix}_n",
        f"{prefix}_avg",
        f"{prefix}_max",
        f"{prefix}_p99",
        f"{prefix}_p95",
        f"{prefix}_p90",
        f"{prefix}_p50",
        f"{prefix}_min",
    ]


def _format_metric_label(key: str) -> str:
    # 图表页/明细页的指标名显示：TTFT/E2E/TPOT 前缀全大写
    show = key
    for p in ("ttft_cot_", "e2e_cot_", "ttft_", "e2e_", "tpot_"):
        if show.startswith(p):
            show = p.upper() + show[len(p) :]
            break
    unit = _metric_unit(key)
    return f"{show}/{unit}" if unit else show


def _split_metric_key(key: str) -> tuple[str, str]:
    """
    将扁平 key 拆成两列展示：
    - 第 1 列：数据项/大类（带单位）
    - 第 2 列：子项（n/avg/... 或 overall/easy/...）
    """
    if key in ("n", "pred_none"):
        return key, ""
    if key.startswith("acc_"):
        return "acc(%)", key[len("acc_") :]
    # latency like ttft_avg / e2e_p99 / ttft_cot_p50 ...
    if "_" in key:
        prefix, sub = key.split("_", 1)
        # cot 前缀：ttft_cot_xxx / e2e_cot_xxx
        if prefix in ("ttft", "e2e") and sub.startswith("cot_"):
            sub2 = sub[len("cot_") :]
            return f"{prefix.upper()}_COT(ms)", sub2
        if prefix in ("ttft", "e2e", "tpot"):
            return f"{prefix.upper()}(ms)", sub
        if key.startswith("input_tokens_"):
            return "INPUT_TOKENS(tok)", key[len("input_tokens_") :]
        if key.startswith("output_tokens_"):
            return "OUTPUT_TOKENS(tok)", key[len("output_tokens_") :]
        if key.startswith("input_tokens_cot_"):
            return "INPUT_TOKENS_COT(tok)", key[len("input_tokens_cot_") :]
        if key.startswith("output_tokens_cot_"):
            return "OUTPUT_TOKENS_COT(tok)", key[len("output_tokens_cot_") :]
    # fallback
    return key, ""


def write_excel(rows: List[Dict[str, Any]], stats: List[BatchStats], out_path: str) -> None:
    """写入 Excel：Accuracy 单独一页；TTFT/E2E/TPOT 另按 domain/sub_domain 分组输出一页；并保留原折线图页（overall）。"""
    wb = openpyxl.Workbook()
    # 默认 sheet 作为 Accuracy
    ws = wb.active
    ws.title = "Accuracy"

    # 统计所有 keys（除 batch）
    all_keys_set = set()
    for r in rows:
        all_keys_set.update(k for k in r.keys() if k != "batch")

    # 若 COT 的 n 全为 0，则隐藏其所有指标与图表页
    def _all_zero(key: str) -> bool:
        vals = [r.get(key) for r in rows]
        return bool(vals) and all((v == 0) for v in vals if isinstance(v, (int, float)))

    hide_ttft_cot = ("ttft_cot_n" in all_keys_set) and _all_zero("ttft_cot_n")
    hide_e2e_cot = ("e2e_cot_n" in all_keys_set) and _all_zero("e2e_cot_n")
    hide_input_tokens_cot = ("input_tokens_cot_n" in all_keys_set) and _all_zero("input_tokens_cot_n")
    hide_output_tokens_cot = ("output_tokens_cot_n" in all_keys_set) and _all_zero("output_tokens_cot_n")

    if hide_ttft_cot:
        all_keys_set = {k for k in all_keys_set if not k.startswith("ttft_cot_")}
    if hide_e2e_cot:
        all_keys_set = {k for k in all_keys_set if not k.startswith("e2e_cot_")}
    if hide_input_tokens_cot:
        all_keys_set = {k for k in all_keys_set if not k.startswith("input_tokens_cot_")}
    if hide_output_tokens_cot:
        all_keys_set = {k for k in all_keys_set if not k.startswith("output_tokens_cot_")}

    # Accuracy 页指标顺序：先放 acc_*（便于直接用主表画折线图且横坐标对齐），再放 n/pred_none
    summary_order = ["acc_overall", "acc_easy", "acc_hard", "acc_short", "acc_medium", "acc_long", "n", "pred_none"]
    # Accuracy 页只展示上述指标，不混入延迟等其它数据
    summary_keys = [k for k in summary_order if k in all_keys_set]

    # Accuracy 页：将“指标/子项”合并为一列；数据列从第 2 列开始
    ws.cell(row=1, column=1, value=SUMMARY_HEADER_LABEL).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).border = BORDER_THIN
    for c, r in enumerate(rows, 2):
        cell = ws.cell(row=1, column=c, value=r.get("batch"))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 转置写入
    for ridx, key in enumerate(summary_keys, 2):
        major, minor = _split_metric_key(key)
        label = f"{major} {minor}".strip() if minor else major
        ws.cell(row=ridx, column=1, value=label).border = BORDER_THIN
        for c, r in enumerate(rows, 2):
            val = r.get(key)
            val = val if val is not None else ""
            cell = ws.cell(row=ridx, column=c, value=val)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = COL_WIDTH_INDEX
    for c in range(2, 2 + len(rows)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = COL_WIDTH_DATA

    # Accuracy 页折线图：只画 acc_* 指标（y 轴为 %）
    acc_chart_keys = ["acc_overall", "acc_easy", "acc_hard", "acc_short", "acc_medium", "acc_long"]
    acc_rows = []
    for k in acc_chart_keys:
        if k in summary_keys:
            acc_rows.append(2 + summary_keys.index(k))
    if acc_rows and len(rows) >= 1:
        chart = LineChart()
        chart.title = "LongBench — Accuracy"
        chart.y_axis.title = "Value (%)"
        chart.x_axis.title = "slice"

        # 直接用主表画图：acc 行在表顶端连续，避免横坐标错位
        r1, r2 = min(acc_rows), max(acc_rows)
        data_acc = Reference(ws, min_col=2, min_row=1, max_col=1 + len(rows), max_row=r2)
        cats = Reference(ws, min_col=1, min_row=r1, max_row=r2)
        chart.add_data(data_acc, titles_from_data=True)
        chart.set_categories(cats)
        vals = []
        for rr in range(r1, r2 + 1):
            for cc in range(2, 2 + len(rows)):
                v = ws.cell(row=rr, column=cc).value
                if isinstance(v, (int, float)):
                    vals.append(float(v))
        if vals:
            vmin, vmax = min(vals), max(vals)
            span = vmax - vmin
            if span <= 0:
                span = abs(vmin) or 1
            margin = max(span * CHART_AXIS_MARGIN_RATIO, span * CHART_AXIS_MARGIN_RATIO_MIN)
            chart.y_axis.scaling.min = vmin - margin
            chart.y_axis.scaling.max = vmax + margin
        chart.width = CHART_WIDTH
        chart.height = CHART_HEIGHT
        ws.add_chart(chart, f"A{(max(acc_rows) + 3)}")

    # 额外：Performance（按 domain/sub_domain 分组）页 —— 4 列 + batch 列
    ws_lat = wb.create_sheet(title="Performance")
    # 列顺序：domain / sub_domain / Metric / stat
    lat_headers = ["domain", "sub_domain", "Metric", "stat"]
    for i, h in enumerate(lat_headers, 1):
        cell = ws_lat.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for c, r in enumerate(rows, 5):
        cell = ws_lat.cell(row=1, column=c, value=r.get("batch"))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # union all groups
    by_batch = [s.latency_by_domain_sub for s in stats]
    group_keys = set()
    for d in by_batch:
        group_keys.update(d.keys())

    def _split_gk(gk: str) -> tuple[str, str]:
        a, b = (gk.split("\t", 1) + [""])[:2]
        return a, b

    stat_order = ["n", "avg", "max", "p99", "p95", "p90", "p50", "min"]
    metrics = ["TTFT", "E2E", "TPOT", "InputTokens", "OutputTokens", "InputTokens_COT", "OutputTokens_COT"]
    metric_unit = {
        "TTFT": "ms",
        "E2E": "ms",
        "TPOT": "ms",
        "InputTokens": "tok",
        "OutputTokens": "tok",
        "InputTokens_COT": "tok",
        "OutputTokens_COT": "tok",
    }
    rowi = 2
    def _sort_key(gk: str) -> tuple[int, str, int, str]:
        dom, sub = _split_gk(gk)
        return (0 if dom == "all" else 1, dom, 0 if sub == "all" else 1, sub)

    for gk in sorted(group_keys, key=_sort_key):
        dom, sub = _split_gk(gk)
        for metric in metrics:
            # 若该 (domain,sub,metric) 在所有 batch 的 n 都为 0，则跳过
            ns = []
            for d in by_batch:
                ns.append((d.get(gk, {}).get(metric, {}) or {}).get("n", 0))
            if all((isinstance(v, (int, float)) and v == 0) for v in ns):
                continue
            for stat_name in stat_order:
                ws_lat.cell(row=rowi, column=1, value=dom).border = BORDER_THIN
                ws_lat.cell(row=rowi, column=2, value=sub).border = BORDER_THIN
                ws_lat.cell(row=rowi, column=3, value=f"{metric}/{metric_unit.get(metric, '')}").border = BORDER_THIN
                ws_lat.cell(row=rowi, column=4, value=stat_name).border = BORDER_THIN
                for bi, d in enumerate(by_batch, 0):
                    val = (d.get(gk, {}).get(metric, {}) or {}).get(stat_name)
                    val = val if val is not None else ""
                    cell = ws_lat.cell(row=rowi, column=5 + bi, value=val)
                    cell.border = BORDER_THIN
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                rowi += 1

    ws_lat.column_dimensions["A"].width = 10
    ws_lat.column_dimensions["B"].width = 28
    ws_lat.column_dimensions["C"].width = 24
    ws_lat.column_dimensions["D"].width = 10
    for c in range(5, 5 + len(rows)):
        ws_lat.column_dimensions[openpyxl.utils.get_column_letter(c)].width = COL_WIDTH_DATA

    # 图表页（overall latency & accuracy）
    def _vals_for(keys: List[str]) -> List[float]:
        vals: List[float] = []
        for r in rows:
            for k in keys:
                v = r.get(k)
                if isinstance(v, (int, float)):
                    vals.append(float(v))
        return vals

    for group_name, metrics in CHART_GROUPS:
        # 避免和 Accuracy 表格页重名产生 Accuracy1；Accuracy 的图可暂不生成
        if group_name == "Accuracy":
            continue
        if hide_ttft_cot and group_name.startswith("COT_TTFT"):
            continue
        if hide_e2e_cot and group_name.startswith("COT_E2E"):
            continue
        # 表格与折线图：都排除 *_n
        table_keys = [
            k
            for k in metrics
            if (k in all_keys_set)
            and (not k.endswith("_n"))
            and any(isinstance(r.get(k), (int, float)) for r in rows)
        ]
        chart_keys = list(table_keys)
        if not table_keys:
            continue
        title = group_name[:SHEET_TITLE_MAX_LEN]
        wsg = wb.create_sheet(title=title)

        # 数据表：A 列为指标名，B.. 为 batch
        start_row = 1
        col_anchor = 1
        wsg.cell(row=start_row, column=col_anchor, value="")
        for c, r in enumerate(rows):
            wsg.cell(row=start_row, column=col_anchor + 1 + c, value=r.get("batch"))
        for i, k in enumerate(table_keys, 1):
            rowi = start_row + i
            wsg.cell(row=rowi, column=col_anchor, value=_format_metric_label(k))
            for c, r in enumerate(rows):
                wsg.cell(row=rowi, column=col_anchor + 1 + c, value=r.get(k))

        # 图
        if not chart_keys:
            continue
        chart = LineChart()
        chart.title = f"LongBench — {group_name}"
        unit = "%" if group_name == "Accuracy" else "ms"
        chart.y_axis.title = f"Value ({unit})"
        chart.x_axis.title = "指标"
        # chart 使用 chart_keys 对应行（假设 *_n 位于表格顶部；若未来顺序变化，这里按索引定位）
        first_idx = table_keys.index(chart_keys[0]) + 1
        last_idx = table_keys.index(chart_keys[-1]) + 1
        data = Reference(
            wsg,
            min_col=col_anchor + 1,
            min_row=start_row,
            max_col=col_anchor + len(rows),
            max_row=start_row + last_idx,
        )
        cats = Reference(wsg, min_col=col_anchor, min_row=start_row + first_idx, max_row=start_row + last_idx)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        vals = _vals_for(chart_keys)
        if vals:
            vmin, vmax = min(vals), max(vals)
            span = vmax - vmin
            if span <= 0:
                span = abs(vmin) or 1
            margin = max(span * CHART_AXIS_MARGIN_RATIO, span * CHART_AXIS_MARGIN_RATIO_MIN)
            chart.y_axis.scaling.min = vmin - margin
            chart.y_axis.scaling.max = vmax + margin
        chart.width = CHART_WIDTH
        chart.height = CHART_HEIGHT
        wsg.add_chart(chart, f"A{start_row + len(table_keys) + 2}")

        wsg.column_dimensions["A"].width = 26
        for c in range(2, 2 + len(rows)):
            wsg.column_dimensions[openpyxl.utils.get_column_letter(c)].width = COL_WIDTH_DATA

    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser(description="Parse LongBench jsonl under batch_* dirs, output Excel summary.")
    p.add_argument("log_dir", help="父目录，其下含以 batch_ 开头的子目录，各子目录下有 long_bench_output.jsonl")
    p.add_argument("--output", default=DEFAULT_OUTPUT_XLSX, help="输出 Excel 文件名（写到 log_dir 下）")
    args = p.parse_args()

    log_dir = os.path.abspath(args.log_dir)
    if not os.path.isdir(log_dir):
        print(f"Error: not a directory: {log_dir}")
        return 1

    batches = collect_batches(log_dir)
    if not batches:
        print(f"No batch_* dirs with {DEFAULT_FILENAME} found under {log_dir}")
        return 1

    all_stats: List[BatchStats] = []
    for b in batches:
        path = os.path.join(log_dir, b, DEFAULT_FILENAME)
        all_stats.append(summarize_one(path, b))

    # 控制台打印关键趋势
    print("LongBench (关键指标 from long_bench_output.jsonl)")
    print("-" * PRINT_SEP_WIDTH)
    print(
        f"{'batch':<16}  {'N':>6}  {'acc_overall(%)':>14}  "
        f"{'ttft_avg(ms)':>12}  {'e2e_avg(ms)':>11}  {'tpot_avg(ms)':>11}  {'out_tok_avg':>11}"
    )
    print("-" * PRINT_SEP_WIDTH)
    for s in all_stats:
        ttft = s.latency_ms["ttft"]
        e2e = s.latency_ms["e2e"]
        tpot = s.latency_ms["tpot"]
        out_tok = s.latency_ms["output_tokens"]
        print(
            f"{s.batch:<16}  {s.n:>6}  {_fmt(s.accuracy.get('overall'), 1):>14}  "
            f"{_fmt(ttft.get('avg'), 0):>12}  {_fmt(e2e.get('avg'), 0):>11}  {_fmt(tpot.get('avg'), 2):>11}  {_fmt(out_tok.get('avg'), 1):>11}"
        )
    print("-" * PRINT_SEP_WIDTH)

    rows = [_flatten_stats(s) for s in all_stats]
    out_xlsx = os.path.join(log_dir, args.output)
    write_excel(rows, all_stats, out_xlsx)
    print(f"Excel saved: {out_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

