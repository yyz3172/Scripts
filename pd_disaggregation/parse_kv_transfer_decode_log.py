#!/usr/bin/env python3
"""
从各 batch 的 decode.log 中解析 Prefill→Decode 的 KV Cache 传输耗时，
输出 Excel：汇总表 + 按 Batch 对比的折线图。

解析规则：匹配 "KV cache transfer for request ... took X.XX ms (N groups, M blocks)"
按请求聚合时取两卡（Worker_TP0/TP1）的最大值作为该请求的传输耗时。

用法：
  python parse_kv_transfer_decode_log.py <log_dir> [--output <文件名.xlsx>]
  # log_dir: 包含 batch_50, batch_60 等子目录的父目录（其下各有 decode.log），Excel 输出到该目录
  # --output: Excel 文件名，默认为 kv_transfer_summary.xlsx
"""

import re
import os
import argparse
from collections import defaultdict

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# KV 传输日志正则（Decode 侧 MooncakeConnector）
KV_TRANSFER_PATTERN = re.compile(
    r"KV cache transfer for request ([^\s]+) took ([\d.]+) ms \((\d+) groups, (\d+) blocks\)"
)

# Excel 表头样式
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def parse_decode_log(decode_log_path):
    """
    解析单个 decode.log，返回按 request_id 聚合的 (transfer_ms_list, blocks_list)。
    每个 request 可能有多行（TP0/TP1），取 max(ms) 作为该请求的传输时间。
    """
    by_req = defaultdict(list)
    with open(decode_log_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            m = KV_TRANSFER_PATTERN.search(line)
            if m:
                req_id, ms, grp, blk = m.group(1), float(m.group(2)), int(m.group(3)), int(m.group(4))
                by_req[req_id].append((ms, blk))
    times = []
    blocks = []
    for req_id, vals in by_req.items():
        ms_vals = [v[0] for v in vals]
        blk_vals = [v[1] for v in vals]
        times.append(max(ms_vals))
        blocks.append(max(blk_vals) if blk_vals else 0)
    return times, blocks


def stats(values):
    """计算 min, avg, p50, p90, p99, max。"""
    if not values:
        return None
    n = len(values)
    s = sorted(values)
    return {
        "count": n,
        "min": min(values),
        "avg": sum(values) / n,
        "p50": s[n // 2],
        "p90": s[int(n * 0.90)] if n > 10 else s[-1],
        "p99": s[int(n * 0.99)] if n > 100 else s[-1],
        "max": max(values),
    }


def collect_batches(log_dir):
    """在 log_dir 下查找所有 batch_* 子目录（含 decode.log），按并发数排序。"""
    batches = []
    for name in os.listdir(log_dir):
        if name.startswith("batch_") and os.path.isdir(os.path.join(log_dir, name)):
            decode_path = os.path.join(log_dir, name, "decode.log")
            if os.path.isfile(decode_path):
                try:
                    concurrency = int(name.replace("batch_", ""))
                    batches.append((name, concurrency))
                except ValueError:
                    continue
    batches.sort(key=lambda x: x[1])
    return batches


def write_excel(results, out_path):
    """写入 Excel：汇总表 + 折线图。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总"

    headers = ["batch", "concurrency", "count", "min_ms", "avg_ms", "p50_ms", "p90_ms", "p99_ms", "max_ms", "avg_blocks", "max_blocks"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
    for r, row in enumerate(results, 2):
        for c, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=r, column=c, value=val if val is not None else "")
            cell.border = BORDER_THIN
    n_results = len(results)
    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 12

    # 折线图数据：横轴=指标(avg_ms, p50_ms, p99_ms, max_ms)，每条线=一个 batch
    valid = [x for x in results if x.get("count")]
    if valid:
        chart_start_row = n_results + 3
        n_valid = len(valid)
        metric_names = ["avg_ms", "p50_ms", "p99_ms", "max_ms"]
        # 第 1 行：空 | batch_50 | batch_60 | ...（系列名）
        ws.cell(row=chart_start_row, column=1, value="")
        for c, r in enumerate(valid, 2):
            ws.cell(row=chart_start_row, column=c, value=r["batch"])
        # 第 2～5 行：指标名 | 各 batch 的该指标值
        for r_idx, m in enumerate(metric_names, 1):
            row = chart_start_row + r_idx
            ws.cell(row=row, column=1, value=m)
            for c, r in enumerate(valid, 2):
                ws.cell(row=row, column=c, value=r[m])
        chart = LineChart()
        chart.title = "KV Cache Transfer (Prefill→Decode)"
        chart.y_axis.title = "ms"
        chart.x_axis.title = "指标"
        data = Reference(ws, min_col=2, min_row=chart_start_row, max_col=1 + n_valid, max_row=chart_start_row + len(metric_names))
        cats = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=chart_start_row + len(metric_names))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 14
        chart.height = 10
        ws.add_chart(chart, f"A{chart_start_row + len(metric_names) + 2}")

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Parse KV transfer from decode.log, output Excel.")
    parser.add_argument("log_dir", help="Parent dir containing batch_50, batch_60, ... with decode.log")
    parser.add_argument("--output", default="kv_transfer_summary.xlsx", help="Output Excel filename (default: kv_transfer_summary.xlsx)")
    args = parser.parse_args()

    log_dir = os.path.abspath(args.log_dir)
    if not os.path.isdir(log_dir):
        print(f"Error: not a directory: {log_dir}")
        return 1
    out_path = os.path.join(log_dir, args.output)

    batches = collect_batches(log_dir)
    if not batches:
        print(f"No batch_* dirs with decode.log found under {log_dir}")
        return 1

    results = []
    for batch_name, concurrency in batches:
        decode_path = os.path.join(log_dir, batch_name, "decode.log")
        times, blocks = parse_decode_log(decode_path)
        if not times:
            results.append({
                "batch": batch_name,
                "concurrency": concurrency,
                "count": 0,
                "min_ms": None, "avg_ms": None, "p50_ms": None, "p90_ms": None, "p99_ms": None, "max_ms": None,
                "avg_blocks": None, "max_blocks": None,
            })
            continue
        st = stats(times)
        results.append({
            "batch": batch_name,
            "concurrency": concurrency,
            "count": st["count"],
            "min_ms": round(st["min"], 2),
            "avg_ms": round(st["avg"], 2),
            "p50_ms": round(st["p50"], 2),
            "p90_ms": round(st["p90"], 2),
            "p99_ms": round(st["p99"], 2),
            "max_ms": round(st["max"], 2),
            "avg_blocks": round(sum(blocks) / len(blocks), 1),
            "max_blocks": max(blocks),
        })

    print("KV cache transfer (Decode 侧, 按请求取两卡最大值)")
    print("-" * 100)
    print(f"{'batch':<12} {'concurrency':>6} {'count':>6} {'avg_ms':>8} {'p50_ms':>8} {'p99_ms':>8} {'max_ms':>8} {'avg_blk':>8}")
    print("-" * 100)
    for r in results:
        if r["count"] == 0:
            print(f"{r['batch']:<12} {r['concurrency']:>6} {r['count']:>6}   --       --       --       --   --")
        else:
            print(f"{r['batch']:<12} {r['concurrency']:>6} {r['count']:>6} {r['avg_ms']:>8.2f} {r['p50_ms']:>8.2f} {r['p99_ms']:>8.2f} {r['max_ms']:>8.2f} {r['avg_blocks']:>8.1f}")
    print("-" * 100)

    write_excel(results, out_path)
    print(f"Excel saved: {out_path}")
    return 0


if __name__ == "__main__":
    exit(main())
