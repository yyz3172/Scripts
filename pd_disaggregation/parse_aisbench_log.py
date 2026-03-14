#!/usr/bin/env python3
"""
从各 batch 的 aisbench.log 中解析 "Performance Results of task" 相关表格数据，
输出 Excel：转置汇总表（带单位）+ 按指标组分页的折线图（E2EL/TTFT/TPOT/ITL/Throughput 等各一 Sheet）。

解析内容：
  1) Performance Parameters 表：E2EL, TTFT, TPOT, ITL, InputTokens, OutputTokens, OutputTokenThroughput 的 Avg/Min/Max/Median/P75/P90/P99/N
  2) Common Metric 表：Benchmark Duration, Total Requests, Request Throughput, 各类 Token Throughput 等

输出说明：
  - 汇总页：转置表，第 1 列为指标名、第 2 列起为各 batch，单元格优先显示带单位字符串（如 "61714.2 ms"）
  - 折线图：每组指标一个 Sheet，横轴为指标、每条线为一个 batch；纵轴按数据范围缩放（不强制从 0 开始），Y 轴标题带单位

配置：表格样式、列宽、折线图尺寸、分组与单位等均在文件开头「配置」区，可按需修改。

用法：
  python parse_aisbench_log.py <log_dir> [--output <文件名.xlsx>]
  log_dir: 包含 batch_50, batch_60 等子目录的父目录（其下各有 aisbench.log），Excel 输出到该目录
  --output: 输出 Excel 文件名，默认见配置 DEFAULT_OUTPUT_FILENAME
"""

import re
import os
import argparse

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side

# -----------------------------------------------------------------------------
# 配置（可按需修改）
# -----------------------------------------------------------------------------
# 表头样式
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
# 折线图尺寸（宽、高）
CHART_WIDTH = 40
CHART_HEIGHT = 20
# 汇总表：Sheet 名、首列表头文案、首行数据指标（不含 batch，因表头列已是 batch 名）
SUMMARY_SHEET_TITLE = "汇总"
SUMMARY_HEADER_LABEL = "指标"
SUMMARY_FIRST_ROW_HEADERS = ["concurrency"]
COL_WIDTH_INDEX = 24      # 汇总表指标列（A 列）宽
COL_WIDTH_DATA = 14       # 汇总表与图表页数据列宽
# 图表页：指标列宽、纵轴边距比例、Excel 表名最大长度
CHART_SHEET_COL_WIDTH_INDEX = 22
CHART_AXIS_MARGIN_RATIO = 0.05
CHART_AXIS_MARGIN_RATIO_MIN = 0.02
SHEET_TITLE_MAX_LEN = 31
# 解析：去除 ANSI 转义的正则、列分隔符（| 与 Unicode │）
ANSI_ESCAPE = re.compile(r"\x1b\[[0-9;]*m")
VERTICAL_BAR = re.compile(r"[|\u2502]")
# Performance Parameters 表：指标名与统计列名（与日志列顺序一致）
PERF_PARAM_NAMES = ("E2EL", "TTFT", "TPOT", "ITL", "InputTokens", "OutputTokens", "OutputTokenThroughput")
PERF_COL_NAMES = ("Avg", "Min", "Max", "Median", "P75", "P90", "P99")
# 折线图分组：每组一个 Sheet，列表项为 (Sheet 名, 该组指标 key 列表)
CHART_GROUPS = [
    ("E2EL", [f"E2EL_{c}" for c in PERF_COL_NAMES]),
    ("TTFT", [f"TTFT_{c}" for c in PERF_COL_NAMES]),
    ("TPOT", [f"TPOT_{c}" for c in PERF_COL_NAMES]),
    ("ITL", [f"ITL_{c}" for c in PERF_COL_NAMES]),
    ("InputTokens", [f"InputTokens_{c}" for c in PERF_COL_NAMES]),
    ("OutputTokens", [f"OutputTokens_{c}" for c in PERF_COL_NAMES]),
    ("OutputTokenThroughput", [f"OutputTokenThroughput_{c}" for c in PERF_COL_NAMES]),
    ("Throughput", ["Request_Throughput", "Prefill_Token_Throughput", "Output_Token_Throughput", "Total_Token_Throughput"]),
]
# 各折线图组对应的 Y 轴单位（用于 Y 轴标题，如 "Value (ms)"）
CHART_GROUP_UNITS = {
    "E2EL": "ms",
    "TTFT": "ms",
    "TPOT": "ms",
    "ITL": "ms",
    "InputTokens": "",
    "OutputTokens": "",
    "OutputTokenThroughput": "token/s",
    "Throughput": "",
}
# 仅写入汇总表、不参与折线图分组的键（当前未用于逻辑，可作扩展参考）
TABLE_ONLY_KEYS = frozenset({
    "Benchmark_Duration", "Total_Requests", "Failed_Requests", "Success_Requests",
    "Concurrency", "Max_Concurrency", "N",
})
# 默认输出 Excel 文件名、控制台打印分隔线宽度
DEFAULT_OUTPUT_FILENAME = "aisbench_performance_summary.xlsx"
PRINT_SEP_WIDTH = 80
# -----------------------------------------------------------------------------


def _parse_value(s):
    """从 '61714.2 ms' / '0.8018 req/s' 等中提取数值，无法解析则返回 None。"""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    m = re.match(r"([\d.]+)", s)
    return float(m.group(1)) if m else None


def _row_cells(line):
    """表格行先去 ANSI 转义，再按 | 或 │ 分割并 strip，返回单元格列表。"""
    line = ANSI_ESCAPE.sub("", line)
    parts = VERTICAL_BAR.split(line)
    return [c.strip() for c in parts]


def parse_aisbench_log(aisbench_log_path):
    """
    解析单个 aisbench.log，提取 Performance Results 中两个表格。
    返回 (out, disp)：out 为数值 dict（用于作图），disp 为带单位的显示串 dict（用于表格）。
    """
    with open(aisbench_log_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    out = {}
    disp = {}  # 同 key，值为原始单元格字符串（含单位）
    in_perf_table = False
    in_common_table = False

    for line in lines:
        line_plain = ANSI_ESCAPE.sub("", line)
        if "Performance Results of task" in line_plain:
            in_perf_table = False
            in_common_table = False
            continue
        cells = _row_cells(line)
        if len(cells) < 3:
            if "╘" in line_plain:
                if in_perf_table:
                    in_perf_table = False
                if in_common_table:
                    in_common_table = False
            continue
        first = (cells[1] if len(cells) > 1 else "") or (cells[0] if cells else "")
        second = (cells[2] if len(cells) > 2 else "") or (cells[1] if len(cells) > 1 else "")
        if "Performance Parameters" in first and "Stage" in second:
            in_perf_table = True
            in_common_table = False
            continue
        if "Common Metric" in first and "Stage" in second:
            in_common_table = True
            in_perf_table = False
            continue
        name = (cells[1] or cells[0] or "").strip()
        if in_perf_table and name and name not in ("Stage", "Performance Parameters"):
            if name in PERF_PARAM_NAMES:
                idx = 3
                for col in PERF_COL_NAMES:
                    if idx < len(cells):
                        raw = (cells[idx] or "").strip()
                        out[f"{name}_{col}"] = _parse_value(raw)
                        disp[f"{name}_{col}"] = raw if raw else ""
                    idx += 1
                if len(cells) > 10:
                    try:
                        out[f"{name}_N"] = int(_parse_value(cells[10]) or 0)
                        disp[f"{name}_N"] = (cells[10] or "").strip()
                    except (ValueError, TypeError):
                        out[f"{name}_N"] = None
                        disp[f"{name}_N"] = ""
        if in_common_table and name and name not in ("Common Metric", "Stage"):
            val = _parse_value(cells[3]) if len(cells) > 3 else _parse_value(cells[2]) if len(cells) > 2 else None
            raw = (cells[3] if len(cells) > 3 else cells[2] if len(cells) > 2 else "" or "").strip()
            key = name.replace(" ", "_")
            out[key] = val
            disp[key] = raw
    return out, disp


def collect_batches(log_dir):
    """在 log_dir 下查找所有 batch_* 子目录（含 aisbench.log），按并发数排序。"""
    batches = []
    for name in os.listdir(log_dir):
        if name.startswith("batch_") and os.path.isdir(os.path.join(log_dir, name)):
            path = os.path.join(log_dir, name, "aisbench.log")
            if os.path.isfile(path):
                try:
                    concurrency = int(name.replace("batch_", ""))
                    batches.append((name, concurrency))
                except ValueError:
                    continue
    batches.sort(key=lambda x: x[1])
    return batches


def write_excel(results, out_path, all_keys, results_disp=None):
    """
    写入 Excel：汇总页（转置表、带单位）+ 按 CHART_GROUPS 每组一个 Sheet 的折线图。
    折线图纵轴按数据范围缩放、Y 轴标题带单位。配置见文件开头。
    """
    if results_disp is None:
        results_disp = [{}] * len(results)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET_TITLE

    # 汇总表转置：第1列=指标名，第2列起=各 batch（表头已含 batch 名，不再单独一行 batch）
    row_headers = list(SUMMARY_FIRST_ROW_HEADERS) + all_keys
    n_results = len(results)
    ws.cell(row=1, column=1, value=SUMMARY_HEADER_LABEL).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).border = BORDER_THIN
    for c, r in enumerate(results, 2):
        cell = ws.cell(row=1, column=c, value=r.get("batch"))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
    for r_idx, key in enumerate(row_headers, 2):
        ws.cell(row=r_idx, column=1, value=key).border = BORDER_THIN
        for c, row in enumerate(results, 2):
            # 优先显示带单位的字符串，否则用数值
            d = results_disp[c - 2] if c - 2 < len(results_disp) else {}
            if key in d and (d[key] or "").strip():
                val = d[key].strip()
            else:
                val = row.get(key)
                val = val if val is not None else ""
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.border = BORDER_THIN
    ws.column_dimensions["A"].width = COL_WIDTH_INDEX
    for c in range(2, 2 + n_results):
        try:
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = COL_WIDTH_DATA
        except Exception:
            pass

    valid = [x for x in results if x.get("batch")]
    if not valid:
        wb.save(out_path)
        return

    def has_data(keys):
        return [k for k in keys if k in all_keys and any(r.get(k) is not None for r in valid)]

    n_valid = len(valid)
    for group_name, candidate_keys in CHART_GROUPS:
        chart_metrics = has_data(candidate_keys)
        if not chart_metrics:
            continue
        sheet_title = group_name[:SHEET_TITLE_MAX_LEN]
        ws_group = wb.create_sheet(title=sheet_title)
        chart_start_row = 1
        col_anchor = 1
        ws_group.cell(row=chart_start_row, column=col_anchor, value="")
        for c, r in enumerate(valid):
            ws_group.cell(row=chart_start_row, column=col_anchor + 1 + c, value=r["batch"])
        for r_idx, m in enumerate(chart_metrics, 1):
            row = chart_start_row + r_idx
            ws_group.cell(row=row, column=col_anchor, value=m)
            for c, r in enumerate(valid):
                ws_group.cell(row=row, column=col_anchor + 1 + c, value=r.get(m))
        # 收集该图数据范围，用于纵轴缩放（不强制从 0 开始）
        vals = []
        for r in valid:
            for m in chart_metrics:
                v = r.get(m)
                if v is not None and isinstance(v, (int, float)):
                    vals.append(v)
        chart = LineChart()
        chart.title = f"AISBench — {group_name}"
        unit = CHART_GROUP_UNITS.get(group_name, "")
        chart.y_axis.title = f"Value ({unit})" if unit else "Value"
        chart.x_axis.title = "指标"
        data = Reference(ws_group, min_col=col_anchor + 1, min_row=chart_start_row, max_col=col_anchor + n_valid, max_row=chart_start_row + len(chart_metrics))
        cats = Reference(ws_group, min_col=col_anchor, min_row=chart_start_row + 1, max_row=chart_start_row + len(chart_metrics))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        if vals:
            vmin, vmax = min(vals), max(vals)
            span = vmax - vmin
            if span <= 0:
                span = abs(vmin) or 1
            margin = max(span * CHART_AXIS_MARGIN_RATIO, span * CHART_AXIS_MARGIN_RATIO_MIN)
            chart.y_axis.scaling.min = vmin - margin
            chart.y_axis.scaling.max = vmax + margin
        chart.width = CHART_WIDTH
        chart.height = CHART_HEIGHT
        ws_group.add_chart(chart, f"A{chart_start_row + len(chart_metrics) + 2}")
        ws_group.column_dimensions["A"].width = CHART_SHEET_COL_WIDTH_INDEX
        for c in range(2, 2 + n_valid):
            try:
                ws_group.column_dimensions[openpyxl.utils.get_column_letter(c)].width = COL_WIDTH_DATA
            except Exception:
                pass

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="Parse AISBench Performance Results from aisbench.log, output Excel.")
    parser.add_argument("log_dir", help="Parent dir containing batch_50, batch_60, ... with aisbench.log")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_FILENAME, help="Output Excel filename")
    args = parser.parse_args()

    log_dir = os.path.abspath(args.log_dir)
    if not os.path.isdir(log_dir):
        print(f"Error: not a directory: {log_dir}")
        return 1
    out_path = os.path.join(log_dir, args.output)

    batches = collect_batches(log_dir)
    if not batches:
        print(f"No batch_* dirs with aisbench.log found under {log_dir}")
        return 1

    results = []
    results_disp = []
    all_keys_set = set()
    for batch_name, concurrency in batches:
        path = os.path.join(log_dir, batch_name, "aisbench.log")
        row, disp = parse_aisbench_log(path)
        row["batch"] = batch_name
        row["concurrency"] = concurrency
        results.append(row)
        results_disp.append(disp)
        all_keys_set.update(k for k in row if k not in ("batch", "concurrency"))

    # 列顺序：先 batch/concurrency，再按 key 排序，保证各 batch 列一致
    all_keys = sorted(all_keys_set)

    print("AISBench Performance Results (from aisbench.log)")
    print("-" * PRINT_SEP_WIDTH)
    print(f"{'batch':<12} {'concurrency':>6}  E2EL_Avg   TTFT_Avg  ReqThroughput  PrefillTok/s  OutTok/s")
    print("-" * PRINT_SEP_WIDTH)
    for r in results:
        e2el = r.get("E2EL_Avg")
        ttft = r.get("TTFT_Avg")
        rt = r.get("Request_Throughput")
        pt = r.get("Prefill_Token_Throughput")
        ot = r.get("Output_Token_Throughput")
        e2el_s = f"{e2el:.0f}" if e2el is not None else "--"
        ttft_s = f"{ttft:.0f}" if ttft is not None else "--"
        rt_s = f"{rt:.3f}" if rt is not None else "--"
        pt_s = f"{pt:.1f}" if pt is not None else "--"
        ot_s = f"{ot:.1f}" if ot is not None else "--"
        print(f"{r['batch']:<12} {r['concurrency']:>6}  {e2el_s:>8}   {ttft_s:>8}  {rt_s:>14}  {pt_s:>12}  {ot_s:>8}")
    print("-" * PRINT_SEP_WIDTH)

    write_excel(results, out_path, all_keys, results_disp)
    print(f"Excel saved: {out_path}")
    return 0


if __name__ == "__main__":
    exit(main())
