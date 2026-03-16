#!/usr/bin/env python3
"""
从各 batch_* 子目录的 prefill.log / decode.log 中解析「吞吐与并发」指标
（Avg prompt/generation throughput、Running/Waiting reqs、KV cache/prefix cache 等），
生成多并发汇总表 Excel：单 sheet，列含 prefill_batch_* 与 decode_batch_*（每个并发档位一组列）。

日志行示例（vLLM 周期性打印的吞吐与并发状态）：
  Engine 000: Avg prompt throughput: 2304.0 tokens/s, Avg generation throughput: 227.6 tokens/s, Running: 30 reqs, Waiting: 0 reqs, ...

用法：
  python throughput_concurrency_sweep_to_excel.py <log_dir> [--output <文件名.xlsx>]
  # log_dir: 如 log/sharegpt_200_yyz_260313/，其下有 batch_30、batch_40 等子目录，各含 prefill.log、decode.log
"""

import re
import os
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 去掉 ANSI 转义（如 [0;36m(APIServer pid=115079)[0;0m）
ANSI_STRIP = re.compile(r"\x1b\[[0-9;]*m")

# 吞吐与并发指标行：时间戳 + 各数值（Engine 000 为 vLLM 日志原文）
THROUGHPUT_CONCURRENCY_LINE_PATTERN = re.compile(
    r"INFO\s+(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+.*?Engine\s+000:\s+"
    r"Avg prompt throughput:\s+([\d.]+)\s+tokens/s,\s+"
    r"Avg generation throughput:\s+([\d.]+)\s+tokens/s,\s+"
    r"Running:\s+(\d+)\s+reqs,\s+Waiting:\s+(\d+)\s+reqs,\s+"
    r"GPU KV cache usage:\s+([\d.]+)%,\s+"
    r"Prefix cache hit rate:\s+([\d.]+)%,\s+"
    r"External prefix cache hit rate:\s+([\d.]+)%"
)

# 指标名（与上面捕获组顺序一致，不含 timestamp）
METRIC_NAMES = [
    "Avg prompt throughput",
    "Avg generation throughput",
    "Running",
    "Waiting",
    "GPU KV cache usage",
    "Prefix cache hit rate",
    "External prefix cache hit rate",
]

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def strip_ansi(line: str) -> str:
    return ANSI_STRIP.sub("", line)


def parse_log_file(log_path: Path) -> list[dict]:
    """解析单个 log 文件，返回按行顺序的字典列表，每项含 timestamp + 各 metric。"""
    rows = []
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            plain = strip_ansi(line)
            m = THROUGHPUT_CONCURRENCY_LINE_PATTERN.search(plain)
            if not m:
                continue
            ts = m.group(1)
            vals = [m.group(i) for i in range(2, 9)]
            row = {"timestamp": ts}
            for name, v in zip(METRIC_NAMES, vals):
                # 数值列：前 5 个为数字，后 3 个为百分比（仍按数字存）
                try:
                    row[name] = float(v)
                except ValueError:
                    row[name] = v
            rows.append(row)
    return rows


def collect_batches(log_dir: Path) -> list[str]:
    """收集 log_dir 下所有 batch_* 目录名，按数字排序。"""
    batches = []
    for p in log_dir.iterdir():
        if p.is_dir() and p.name.startswith("batch_"):
            try:
                n = int(p.name.split("_")[1])
                batches.append((n, p.name))
            except (IndexError, ValueError):
                continue
    batches.sort(key=lambda x: x[0])
    return [b[1] for b in batches]


def build_single_sheet(wb, sheet_name: str, prefill_data: dict, decode_data: dict):
    """
    单 sheet：先所有 prefill_batch_* 列，再所有 decode_batch_* 列。
    行按索引对齐，不足的留空。
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    batches = sorted(
        set(prefill_data.keys()) | set(decode_data.keys()),
        key=lambda b: int(b.split("_")[1]),
    )
    if not batches:
        ws["A1"] = "无数据"
        return

    # 表头：先 prefill 块，再 decode 块；batch_30 -> BS=30
    def bs_label(batch: str) -> str:
        return f"BS={batch.split('_')[1]}"

    col_idx = 1
    for batch in batches:
        lb = bs_label(batch)
        ws.cell(1, col_idx, f"prefill_{lb}_timestamp")
        col_idx += 1
        for name in METRIC_NAMES:
            ws.cell(1, col_idx, f"prefill_{lb}_{name}")
            col_idx += 1
    for batch in batches:
        lb = bs_label(batch)
        ws.cell(1, col_idx, f"decode_{lb}_timestamp")
        col_idx += 1
        for name in METRIC_NAMES:
            ws.cell(1, col_idx, f"decode_{lb}_{name}")
            col_idx += 1
    total_cols = col_idx - 1

    for c in range(1, total_cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN

    max_rows = max(
        max(len(prefill_data.get(b, [])) for b in batches),
        max(len(decode_data.get(b, [])) for b in batches),
    )
    cols_per_batch = 1 + len(METRIC_NAMES)

    for r in range(max_rows):
        col_idx = 1
        # prefill 块
        for batch in batches:
            rows_list = prefill_data.get(batch, [])
            if r < len(rows_list):
                row = rows_list[r]
                ws.cell(r + 2, col_idx, row["timestamp"])
                col_idx += 1
                for name in METRIC_NAMES:
                    ws.cell(r + 2, col_idx, row.get(name, ""))
                    col_idx += 1
            else:
                col_idx += cols_per_batch
        # decode 块
        for batch in batches:
            rows_list = decode_data.get(batch, [])
            if r < len(rows_list):
                row = rows_list[r]
                ws.cell(r + 2, col_idx, row["timestamp"])
                col_idx += 1
                for name in METRIC_NAMES:
                    ws.cell(r + 2, col_idx, row.get(name, ""))
                    col_idx += 1
            else:
                col_idx += cols_per_batch
        for c in range(1, total_cols + 1):
            ws.cell(r + 2, c).border = BORDER_THIN

    for c in range(1, total_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 20


def main():
    parser = argparse.ArgumentParser(description="从各 batch 的 prefill/decode.log 解析吞吐与并发指标，生成多并发汇总 Excel")
    parser.add_argument("log_dir", type=str, help="包含 batch_30、batch_40 等子目录的路径")
    parser.add_argument("--output", "-o", type=str, default="throughput_concurrency_sweep.xlsx", help="输出 Excel 文件名")
    args = parser.parse_args()

    log_dir = Path(args.log_dir)
    if not log_dir.is_dir():
        raise SystemExit(f"目录不存在: {log_dir}")

    batches = collect_batches(log_dir)
    if not batches:
        raise SystemExit(f"未找到 batch_* 子目录: {log_dir}")

    # 按 log 类型收集：prefill / decode
    prefill_data = {}
    decode_data = {}

    for batch in batches:
        batch_path = log_dir / batch
        prefill_log = batch_path / "prefill.log"
        decode_log = batch_path / "decode.log"

        if prefill_log.exists():
            prefill_data[batch] = parse_log_file(prefill_log)
        else:
            prefill_data[batch] = []

        if decode_log.exists():
            decode_data[batch] = parse_log_file(decode_log)
        else:
            decode_data[batch] = []

    out_path = log_dir / args.output if not os.path.isabs(args.output) else Path(args.output)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_single_sheet(wb, "throughput_concurrency_sweep", prefill_data, decode_data)

    wb.save(out_path)
    print(f"已写入: {out_path}")

if __name__ == "__main__":
    main()
