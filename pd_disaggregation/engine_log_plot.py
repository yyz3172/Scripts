"""
Engine 日志解析与绘图
从 prefill.log / decode.log 中解析 vLLM Engine 指标行，输出 Excel：原始数据 + 折线图。

单目录模式：输入目录下直接含 prefill.log / decode.log 时，在该目录生成 engine_log_analysis.xlsx
（原始数据表 + Throughput / Running_Waiting / Cache 三类图，每类「原始 + 采样」两张图并排）。

批量模式：输入目录下无日志、仅有子目录（如 batch_50, batch_60）时，每个子目录生成各自的
engine_log_analysis.xlsx，并在父目录生成 engine_log_sweep.xlsx（7 指标 × prefill+decode，
每 sheet 横轴 index、多子目录折线对比）。

用法：
  python engine_log_plot.py <log_dir> [--output 文件名.xlsx]
  # --output 仅指定文件名，输出路径始终在 prefill.log 同级；批量时汇总表固定为 engine_log_sweep.xlsx
"""

import re
import os
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter


# ── 配置（可在此集中修改）────────────────────────────────────────────────────
# 解析
ENGINE_LINE_PATTERN = re.compile(
    r"(\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})"
    r".*?"
    r"Engine\s+\d+:"
    r"\s*Avg prompt throughput:\s*([\d.]+)\s*tokens/s,"
    r"\s*Avg generation throughput:\s*([\d.]+)\s*tokens/s,"
    r"\s*Running:\s*(\d+)\s*reqs,"
    r"\s*Waiting:\s*(\d+)\s*reqs,"
    r"\s*GPU KV cache usage:\s*([\d.]+)%,"
    r"\s*Prefix cache hit rate:\s*([\d.]+)%,"
    r"\s*External prefix cache hit rate:\s*([\d.]+)%",
    re.IGNORECASE,
)
COLUMNS = [
    "avg_prompt_throughput",
    "avg_generation_throughput",
    "running",
    "waiting",
    "gpu_kv_cache_usage_pct",
    "prefix_cache_hit_rate_pct",
    "external_prefix_cache_hit_rate_pct",
]
FORWARD_FILL_COLUMNS = {"gpu_kv_cache_usage_pct", "prefix_cache_hit_rate_pct", "external_prefix_cache_hit_rate_pct"}
TIME_ALIGN_TOLERANCE_SEC = 5  # 时间对齐容差（秒），该范围内 prefill/decode 视为同一时刻

# Excel 样式与列宽
FONT_NAME = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
BORDER = Border(
    left=Side(border_style="thin", color="BFBFBF"),
    right=Side(border_style="thin", color="BFBFBF"),
    top=Side(border_style="thin", color="BFBFBF"),
    bottom=Side(border_style="thin", color="BFBFBF"),
)
CENTER = Alignment(horizontal="center", vertical="center")
COL_WIDTH_TIME = 20
COL_WIDTH_INDEX = 10
COL_WIDTH_DATA = 16
RAW_HEADERS = ["index", "timestamp"] + COLUMNS

# 图 sheet 布局与采样
CHART_SHEET_DATA_START_ROW = 55
CHART2_ANCHOR = "O1"
MAX_CHART_POINTS = 100
LINE_WIDTH_EMU = 28000
# 折线图尺寸（宽×高，单位与 Excel 列宽一致）
CHART_WIDTH = 24
CHART_HEIGHT = 16
CHART_SWEEP_WIDTH = 24
CHART_SWEEP_HEIGHT = 16

# 汇总表（批量场景）指标显示名
METRIC_SWEEP_TITLES = {
    "avg_prompt_throughput": "prompt throughput",
    "avg_generation_throughput": "generation throughput",
    "running": "Running",
    "waiting": "Waiting",
    "gpu_kv_cache_usage_pct": "KV cache usage",
    "prefix_cache_hit_rate_pct": "Prefix hit",
    "external_prefix_cache_hit_rate_pct": "External prefix hit",
}


def parse_engine_line(line):
    """从一行日志中解析出指标字典，不匹配则返回 None。"""
    m = ENGINE_LINE_PATTERN.search(line)
    if not m:
        return None
    return {
        "timestamp": m.group(1).strip(),
        "avg_prompt_throughput": float(m.group(2)),
        "avg_generation_throughput": float(m.group(3)),
        "running": int(m.group(4)),
        "waiting": int(m.group(5)),
        "gpu_kv_cache_usage_pct": float(m.group(6)),
        "prefix_cache_hit_rate_pct": float(m.group(7)),
        "external_prefix_cache_hit_rate_pct": float(m.group(8)),
    }


def load_log_series(log_path):
    """从单个日志文件中按行解析所有 Engine 指标，返回列表 of dict。"""
    rows = []
    if not os.path.isfile(log_path):
        return rows
    with open(log_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            d = parse_engine_line(line)
            if d is not None:
                rows.append(d)
    return rows


def load_prefill_decode(log_dir):
    """从 log_dir 下读取 prefill.log 和 decode.log，返回 (prefill_rows, decode_rows)。"""
    log_dir = os.path.abspath(log_dir)
    prefill_path = os.path.join(log_dir, "prefill.log")
    decode_path = os.path.join(log_dir, "decode.log")
    prefill_rows = load_log_series(prefill_path)
    decode_rows = load_log_series(decode_path)
    return prefill_rows, decode_rows


def _dir_has_logs(path):
    """目录下是否直接包含 prefill.log 或 decode.log（视为单目录模式）。"""
    return os.path.isfile(os.path.join(path, "prefill.log")) or os.path.isfile(os.path.join(path, "decode.log"))


def _parse_timestamp(ts_str):
    """将 'MM-DD HH:MM:SS' 转为可比较的 float（固定年份保证顺序）。"""
    try:
        dt = datetime.strptime(ts_str.strip(), "%m-%d %H:%M:%S")
        return dt.replace(year=2024).timestamp()
    except (ValueError, TypeError):
        return None


def align_by_time(prefill_rows, decode_rows):
    """
    按时间对齐 prefill 与 decode：将时间差在 TIME_ALIGN_TOLERANCE_SEC 内的样本视为同一时刻，
    合并为一行（不再用并集导致行数≈prefill数+decode数）。
    返回 [(time_label, prefill_row_or_None, decode_row_or_None), ...]，按时间排序。
    """
    def with_ts(rows, src):
        out = []
        for r in rows:
            t = _parse_timestamp(r.get("timestamp", ""))
            if t is not None:
                out.append((t, src, r))
        return out

    events = with_ts(prefill_rows, "p") + with_ts(decode_rows, "d")
    if not events:
        return []
    events.sort(key=lambda x: x[0])

    # 按容差聚类：同一段内的 (t, src, row) 归为一簇
    clusters = []
    i = 0
    while i < len(events):
        t0 = events[i][0]
        cluster = [events[i]]
        i += 1
        while i < len(events) and events[i][0] - t0 <= TIME_ALIGN_TOLERANCE_SEC:
            cluster.append(events[i])
            i += 1
        clusters.append(cluster)

    result = []
    for cluster in clusters:
        ts = [x[0] for x in cluster]
        rep_t = sum(ts) / len(ts)
        label = datetime.fromtimestamp(rep_t).strftime("%m-%d %H:%M:%S")
        p_entries = [(x[0], x[2]) for x in cluster if x[1] == "p"]
        d_entries = [(x[0], x[2]) for x in cluster if x[1] == "d"]
        p_row = min(p_entries, key=lambda e: abs(e[0] - rep_t))[1] if p_entries else None
        d_row = min(d_entries, key=lambda e: abs(e[0] - rep_t))[1] if d_entries else None
        result.append((label, p_row, d_row))
    return result


def _set_sheet_column_widths(ws, widths):
    """按列表依次设置各列宽。widths[i] 为第 i+1 列宽度。"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _max_len(prefill_rows, decode_rows):
    return max(len(prefill_rows), len(decode_rows), 1)


def _pad(rows, length, fill=None):
    """将 rows 填充到 length 长度，不足用 fill。"""
    return list(rows) + [fill] * (length - len(rows))


def write_raw_sheet(wb, prefill_rows, decode_rows):
    """写入原始数据：两个 sheet Prefill / Decode。"""
    # Prefill
    ws_prefill = wb.create_sheet("Prefill 原始数据", 0)
    for c, h in enumerate(RAW_HEADERS, 1):
        cell = ws_prefill.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for r, row in enumerate(prefill_rows, 2):
        for col_idx, val in enumerate([r - 2, row.get("timestamp", "")] + [row.get(key) for key in COLUMNS], 1):
            cell = ws_prefill.cell(row=r, column=col_idx, value=val)
            cell.alignment = CENTER
    _set_sheet_column_widths(ws_prefill, [COL_WIDTH_INDEX, COL_WIDTH_TIME] + [COL_WIDTH_DATA] * len(COLUMNS))
    # Decode
    ws_decode = wb.create_sheet("Decode 原始数据", 1)
    for c, h in enumerate(RAW_HEADERS, 1):
        cell = ws_decode.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for r, row in enumerate(decode_rows, 2):
        for col_idx, val in enumerate([r - 2, row.get("timestamp", "")] + [row.get(key) for key in COLUMNS], 1):
            cell = ws_decode.cell(row=r, column=col_idx, value=val)
            cell.alignment = CENTER
    _set_sheet_column_widths(ws_decode, [COL_WIDTH_INDEX, COL_WIDTH_TIME] + [COL_WIDTH_DATA] * len(COLUMNS))


def _downsample_indices(n, max_pts):
    """
    返回用于绘图的索引列表。若 n <= max_pts 则全选；否则均匀采样，保留首尾，共约 max_pts 个点。
    """
    if n <= max_pts:
        return list(range(n))
    step = (n - 1) / (max_pts - 1)
    indices = [0]
    for i in range(1, max_pts - 1):
        indices.append(int(round(i * step)))
    indices.append(n - 1)
    return indices


def _write_data_block_aligned(ws, row0, headers, columns_subset, aligned_rows):
    """
    在 sheet 的 row0 起写入一块表头+数据；aligned_rows = [(time_label, p_row, d_row), ...]。
    仅 FORWARD_FILL_COLUMNS 中的指标在缺侧时做前向填充；其余指标该侧留空。
    """
    n_rows = len(aligned_rows)
    time_col = 1
    data_start_col = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row0, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    last_p = {col: None for col in columns_subset if col in FORWARD_FILL_COLUMNS}
    last_d = {col: None for col in columns_subset if col in FORWARD_FILL_COLUMNS}
    for ri, (time_label, p_row, d_row) in enumerate(aligned_rows):
        sheet_row = row0 + 1 + ri
        cell = ws.cell(row=sheet_row, column=time_col, value=time_label)
        cell.alignment = CENTER
        for ci, col in enumerate(columns_subset):
            prefill_col = data_start_col + ci * 2
            decode_col = data_start_col + ci * 2 + 1
            raw_p = p_row.get(col) if p_row else None
            raw_d = d_row.get(col) if d_row else None
            if col in FORWARD_FILL_COLUMNS:
                if raw_p is not None:
                    last_p[col] = raw_p
                if raw_d is not None:
                    last_d[col] = raw_d
                pv = raw_p if raw_p is not None else last_p.get(col)
                dv = raw_d if raw_d is not None else last_d.get(col)
            else:
                pv = raw_p
                dv = raw_d
            cell_p = ws.cell(row=sheet_row, column=prefill_col, value=pv if pv is not None else "")
            cell_p.alignment = CENTER
            cell_d = ws.cell(row=sheet_row, column=decode_col, value=dv if dv is not None else "")
            cell_d.alignment = CENTER
    return n_rows


def _build_line_chart(ws, row0, n_rows, columns_subset, titles, chart_title, num_metrics, palette):
    """根据已有数据区 (row0 表头, row0+1..row0+n_rows 数据) 建折线图并返回 chart。"""
    data_start_col = 2
    chart = LineChart()
    chart.title = chart_title
    chart.style = 10
    chart.y_axis.title = titles[0] if len(titles) == 1 else "Value"
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT
    cats = Reference(ws, min_col=1, min_row=row0 + 1, max_row=row0 + n_rows)
    chart.set_categories(cats)
    for ci in range(num_metrics):
        prefill_col = data_start_col + ci * 2
        decode_col = data_start_col + ci * 2 + 1
        title_prefill = f"prefill {titles[ci]}"
        title_decode = f"decode {titles[ci]}"
        ref_p = Reference(ws, min_col=prefill_col, min_row=row0, max_row=row0 + n_rows)
        ref_d = Reference(ws, min_col=decode_col, min_row=row0, max_row=row0 + n_rows)
        chart.add_data(ref_p, titles_from_data=True)
        chart.series[-1].tx = SeriesLabel(v=title_prefill)
        chart.series[-1].smooth = True
        chart.add_data(ref_d, titles_from_data=True)
        chart.series[-1].tx = SeriesLabel(v=title_decode)
        chart.series[-1].smooth = True
    _style_series_pairwise(chart, num_metrics, palette)
    return chart


def _chart_data_sheet(wb, prefill_rows, decode_rows, name, columns_subset, titles):
    """
    一个 sheet 两张图：按时间对齐 prefill 与 decode，相近时刻的数据才在同一行对比。
    原始数据图 + 采样图；下方对应两块数据区（第一列为时间）。
    """
    aligned = align_by_time(prefill_rows, decode_rows)
    if not aligned:
        n = _max_len(prefill_rows, decode_rows)
        aligned = [
            (str(i), prefill_rows[i] if i < len(prefill_rows) else None, decode_rows[i] if i < len(decode_rows) else None)
            for i in range(n)
        ]
    n = len(aligned)
    indices_sampled = _downsample_indices(n, MAX_CHART_POINTS)
    aligned_sampled = [aligned[i] for i in indices_sampled]
    n_sampled = len(aligned_sampled)

    headers = ["time"]
    for col in columns_subset:
        headers.append(f"prefill_{col}")
        headers.append(f"decode_{col}")

    ws = wb.create_sheet(name)
    num_metrics = len(columns_subset)
    palette = _chart_palette(num_metrics)

    # 数据区1：按时间对齐的完整数据
    row0_full = CHART_SHEET_DATA_START_ROW
    _write_data_block_aligned(ws, row0_full, headers, columns_subset, aligned)
    n_full = n

    # 数据区2：按时间对齐后的采样
    row0_sampled = row0_full + n_full + 2
    _write_data_block_aligned(ws, row0_sampled, headers, columns_subset, aligned_sampled)

    # 图1：原始数据
    chart1 = _build_line_chart(ws, row0_full, n_full, columns_subset, titles, f"{name} (原始数据)", num_metrics, palette)
    ws.add_chart(chart1, "A1")
    # 图2：采样
    chart2 = _build_line_chart(ws, row0_sampled, n_sampled, columns_subset, titles, f"{name} (采样)", num_metrics, palette)
    ws.add_chart(chart2, CHART2_ANCHOR)
    n_cols = len(headers)
    _set_sheet_column_widths(ws, [COL_WIDTH_TIME] + [COL_WIDTH_DATA] * (n_cols - 1))
    return ws


def _chart_palette(num_metrics):
    """
    按指标数量返回颜色对列表：每指标一色系，(深色, 浅色) 用于 prefill / decode。
    不同指标用不同色系，同指标同色系便于对比。
    """
    # 每行 (深色hex, 浅色hex)：蓝、绿、红、紫、橙
    full = [
        ("1B5E9E", "64B5F6"),   # 蓝
        ("2E7D32", "81C784"),   # 绿
        ("C62828", "E57373"),   # 红
        ("6A1B9A", "B39DDB"),   # 紫
        ("E65100", "FFB74D"),   # 橙
    ]
    return full[:num_metrics]


def _style_series_pairwise(chart, num_metrics, colors):
    """
    按「指标」成对设置样式：同指标同色系（prefill 深色、decode 浅色），全部实线，折线较细。
    series 顺序为 [prefill_m0, decode_m0, prefill_m1, decode_m1, ...]
    """
    for i in range(len(chart.series)):
        metric_idx = i // 2
        is_decode = (i % 2) == 1
        pair = colors[metric_idx] if metric_idx < len(colors) else ("333333", "999999")
        color = pair[1] if is_decode else pair[0]
        line = chart.series[i].graphicalProperties.line
        line.solidFill = color
        line.dashStyle = "solid"
        line.width = LINE_WIDTH_EMU


def write_charts(wb, prefill_rows, decode_rows):
    """创建三张图的数据表并插入折线图。"""
    # 图1: Avg prompt throughput, Avg generation throughput
    _chart_data_sheet(
        wb,
        prefill_rows,
        decode_rows,
        name="图1_Throughput",
        columns_subset=["avg_prompt_throughput", "avg_generation_throughput"],
        titles=["Avg prompt throughput (tokens/s)", "Avg generation throughput (tokens/s)"],
    )
    # 图2: Running, Waiting
    _chart_data_sheet(
        wb,
        prefill_rows,
        decode_rows,
        name="图2_Running_Waiting",
        columns_subset=["running", "waiting"],
        titles=["Running (reqs)", "Waiting (reqs)"],
    )
    # 图3: GPU KV cache usage, Prefix cache hit rate, External prefix cache hit rate
    _chart_data_sheet(
        wb,
        prefill_rows,
        decode_rows,
        name="图3_Cache",
        columns_subset=[
            "gpu_kv_cache_usage_pct",
            "prefix_cache_hit_rate_pct",
            "external_prefix_cache_hit_rate_pct",
        ],
        titles=[
            "GPU KV cache usage (%)",
            "Prefix cache hit rate (%)",
            "External prefix cache hit rate (%)",
        ],
    )


def _process_one_dir(log_dir, out_path, prefill_rows=None, decode_rows=None):
    """处理单个目录并写入 Excel；返回 True 成功，False 无有效数据。若未传入 rows 则从 log_dir 加载。"""
    if prefill_rows is None or decode_rows is None:
        prefill_rows, decode_rows = load_prefill_decode(log_dir)
    if not prefill_rows and not decode_rows:
        return False
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    write_raw_sheet(wb, prefill_rows, decode_rows)
    write_charts(wb, prefill_rows, decode_rows)
    wb.save(out_path)
    return True


def _mean_over_rows(rows, col):
    """对多行数据取某列均值，忽略 None；无有效值时返回 None。"""
    vals = [r.get(col) for r in rows if r.get(col) is not None]
    if not vals:
        return None
    return sum(vals) / len(vals)


def _sanitize_sheet_title(title):
    """Excel sheet 名不能含 \\ / * ? [ ] : ，替换为下划线并截断至 31 字符。"""
    for c in r'\/*?[]:':
        title = title.replace(c, "_")
    return title[:31] if len(title) > 31 else title


def _write_sweep_data_block(ws, row0, batch_series, batch_entries, indices):
    """写入一块汇总数据：index + 各 batch 列，仅 indices 中的行。"""
    n_batches = len(batch_entries)
    ws.cell(row=row0, column=1, value="index")
    for c, (name, _, _) in enumerate(batch_entries, 2):
        ws.cell(row=row0, column=c, value=name)
    for c in range(1, n_batches + 2):
        cell = ws.cell(row=row0, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    for ri, idx in enumerate(indices):
        sheet_row = row0 + 1 + ri
        ws.cell(row=sheet_row, column=1, value=idx)
        for bi, series in enumerate(batch_series):
            val = series[idx] if idx < len(series) else None
            ws.cell(row=sheet_row, column=2 + bi, value=val if val is not None else "")
        for c in range(1, n_batches + 2):
            ws.cell(row=sheet_row, column=c).alignment = CENTER


def _add_sweep_chart(ws, row0, n_rows, batch_entries, title, palette):
    """在 ws 上根据 row0 起的数据区添加一张汇总折线图。"""
    n_batches = len(batch_entries)
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = title
    chart.width = CHART_SWEEP_WIDTH
    chart.height = CHART_SWEEP_HEIGHT
    cats = Reference(ws, min_col=1, min_row=row0 + 1, max_row=row0 + n_rows)
    chart.set_categories(cats)
    for bi in range(n_batches):
        ref = Reference(ws, min_col=2 + bi, min_row=row0, max_row=row0 + n_rows)
        chart.add_data(ref, titles_from_data=True)
        chart.series[-1].tx = SeriesLabel(v=batch_entries[bi][0])
        chart.series[-1].smooth = True
    for i in range(len(chart.series)):
        line = chart.series[i].graphicalProperties.line
        pair = palette[i % len(palette)] if i < len(palette) else ("333333", "999999")
        line.solidFill = pair[0]
        line.dashStyle = "solid"
        line.width = LINE_WIDTH_EMU
    return chart


def build_sweep_excel(log_dir, batch_entries, sweep_filename):
    """
    批量场景：14 个 Sheet = 7 个指标 × (prefill + decode)。每个 sheet：横轴 = index，纵轴 = 指标值，
    且含两个图：原始数据图 + 采样（MAX_CHART_POINTS）图。
    """
    if not batch_entries:
        return
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for col in COLUMNS:
        title_base = METRIC_SWEEP_TITLES.get(col, col)
        for source in ("prefill", "decode"):
            sheet_title = _sanitize_sheet_title(f"{title_base}_{source}")
            ws = wb.create_sheet(sheet_title, len(wb.sheetnames))
            batch_series = []
            for _name, prefill_rows, decode_rows in batch_entries:
                rows = prefill_rows if source == "prefill" else decode_rows
                batch_series.append([r.get(col) for r in rows] if rows else [])
            n_points = max(len(s) for s in batch_series) if batch_series else 0
            if n_points == 0:
                continue
            n_batches = len(batch_entries)
            indices_full = list(range(n_points))
            indices_sampled = _downsample_indices(n_points, MAX_CHART_POINTS)
            n_sampled = len(indices_sampled)

            row0_full = CHART_SHEET_DATA_START_ROW
            _write_sweep_data_block(ws, row0_full, batch_series, batch_entries, indices_full)
            row0_sampled = row0_full + n_points + 2
            _write_sweep_data_block(ws, row0_sampled, batch_series, batch_entries, indices_sampled)

            palette = _chart_palette(n_batches)
            chart1 = _add_sweep_chart(ws, row0_full, n_points, batch_entries, f"{title_base} ({source}) 原始数据", palette)
            ws.add_chart(chart1, "A1")
            chart2 = _add_sweep_chart(ws, row0_sampled, n_sampled, batch_entries, f"{title_base} ({source}) 采样", palette)
            ws.add_chart(chart2, CHART2_ANCHOR)

            _set_sheet_column_widths(ws, [COL_WIDTH_INDEX] + [COL_WIDTH_DATA] * n_batches)

    out_path = os.path.join(log_dir, sweep_filename)
    wb.save(out_path)
    print(f"已保存汇总表: {out_path}")


def process_log_dir(log_dir, output_filename):
    """
    仅两种情形：
    1）log_dir 下直接有 prefill.log / decode.log：在该目录下生成 Excel，返回 1。
    2）log_dir 下无日志，仅有若干直接子目录且子目录下有 prefill.log / decode.log：对每个此类子目录生成 Excel，返回处理数量。
    """
    log_dir = os.path.abspath(log_dir)

    if _dir_has_logs(log_dir):
        out_path = os.path.join(log_dir, output_filename)
        prefill_rows, decode_rows = load_prefill_decode(log_dir)
        if not prefill_rows and not decode_rows:
            print("未解析到任何 Engine 指标行，请确认日志格式。")
            return 0
        _process_one_dir(log_dir, out_path, prefill_rows, decode_rows)
        print(f"[单目录] {log_dir}")
        print(f"  Prefill 解析行数: {len(prefill_rows)}, Decode 解析行数: {len(decode_rows)}")
        print(f"已保存: {out_path}")
        return 1

    # 当前目录无日志，仅遍历直接子目录（一层）
    print(f"[批量] 输出文件名: {output_filename}")
    batch_list = []
    for name in sorted(os.listdir(log_dir)):
        sub = os.path.join(log_dir, name)
        if os.path.isdir(sub) and _dir_has_logs(sub):
            prefill_rows, decode_rows = load_prefill_decode(sub)
            if not prefill_rows and not decode_rows:
                print(f"  [SKIP] {name} (无有效数据)")
                continue
            batch_list.append((name, prefill_rows, decode_rows))
            out_path = os.path.join(sub, output_filename)
            _process_one_dir(sub, out_path, prefill_rows, decode_rows)
            print(f"  [OK] {name} (prefill={len(prefill_rows)}, decode={len(decode_rows)})")
    if not batch_list:
        print(f"错误：{log_dir} 下既无 prefill.log/decode.log，也无含日志的直接子目录。")
        return 0
    batch_list.sort(key=lambda x: x[0])  # 按子目录名排序
    sweep_filename = "engine_log_sweep.xlsx"
    build_sweep_excel(log_dir, batch_list, sweep_filename)
    print(f"批量完成: {len(batch_list)} 个目录已生成 Excel，并已生成汇总表 {sweep_filename}")
    return len(batch_list)


def main():
    parser = argparse.ArgumentParser(
        description="解析 prefill/decode 日志并生成 Excel 报表与折线图。"
        " 两种用法：1）log_dir 下直接含 prefill.log/decode.log；2）log_dir 下仅有若干子目录，子目录含日志。"
    )
    parser.add_argument(
        "log_dir",
        type=str,
        help="情形1：含 prefill.log/decode.log 的目录；情形2：其直接子目录含上述日志的父目录",
    )
    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="输出 Excel 文件名（仅文件名，不含路径）；文件始终生成在 prefill.log 同级目录，默认 engine_log_analysis.xlsx",
    )
    args = parser.parse_args()
    log_dir = os.path.abspath(args.log_dir)
    if not os.path.isdir(log_dir):
        print(f"错误：目录不存在 {log_dir}")
        return 1

    default_filename = "engine_log_analysis.xlsx"
    output_filename = (os.path.basename(args.output).strip() if args.output else default_filename) or default_filename
    if not output_filename.lower().endswith(".xlsx"):
        output_filename += ".xlsx"

    n = process_log_dir(log_dir, output_filename)
    return 0 if n > 0 else 1


if __name__ == "__main__":
    exit(main())
